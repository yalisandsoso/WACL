# -*- coding: UTF-8 -*-

import os
import re
import time

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, borders, colors, PatternFill
from openpyxl.utils import get_column_letter


class Aclmat:
    """
        The format when recoding 'aclfile' to the '.xlsx' file
    """

    def __init__(self, lang='ch'):

        self.lang = lang
        return

    def get_title_map(self):
        """
            The specific contents in the advanced security settings of the folder.
        """
        _map = {}
        if self.lang == 'ch':
            _map = {
                'path': '查询路径',
                'domain': '所属组/域',
                'user': '用户名称',
                'fullAccessMask': '完整的权限掩码',
                'accessMask': '不包含继承关系的权限掩码',
                'inherit': '继承关系',
                'parentInherit': '从父对象继承',
                'propagateInherit': '传播继承',
                'inheritRight': '权限的应用场景',
                'isAllow': '是否允许的权限',
                'fullControl': '完全控制',
                'readData_listDir': '列出文件夹/读取数据',
                'readAttr': '读取属性',
                'readExtAttr': '读取扩展属性',
                'readPermiss': '读取权限',
                'execute_traverse': '遍历文件夹/执行文件',
                'writeData_addFile': '创建文件/写入数据',
                'appendData_addSubdir': '创建文件夹/附加数据',
                'writeAttr': '写入属性',
                'writeExtAttr': '写入扩展属性',
                'delete': '删除',
                'deleteChild': '删除子文件夹及文件',
                'changePermiss': '更改权限',
                'takeOwner': '取得所有权',
                'sync': '同步'
            }

        return _map

    def get_per_map(self):
        """
            The specific contents in the advanced security settings of the folder.
        """
        _map = {}
        if self.lang == 'ch':
            _map = {
                'isAllow': '是否允许的权限',
                'fullControl': '完全控制',
                'readData_listDir': '列出文件夹/读取数据',
                'readAttr': '读取属性',
                'readExtAttr': '读取扩展属性',
                'readPermiss': '读取权限',
                'execute_traverse': '遍历文件夹/执行文件',
                'writeData_addFile': '创建文件/写入数据',
                'appendData_addSubdir': '创建文件夹/附加数据',
                'writeAttr': '写入属性',
                'writeExtAttr': '写入扩展属性',
                'delete': '删除',
                'deleteChild': '删除子文件夹及文件',
                'changePermiss': '更改权限',
                'takeOwner': '取得所有权',
                'sync': '同步'
            }

        return _map

    def get_inheritRight_int_map(self):
        """
            The permissions applied.
        """
        _map = {}
        if self.lang == 'ch':
            _map = {
                0: '只有该文件夹',
                1: '只有子文件夹',
                2: '只有文件',
                3: '此文件夹和子文件夹',
                4: '此文件夹和文件',
                5: '仅子文件夹和文件',
                6: '此文件夹、子文件夹和文件'
            }

        return _map

    def get_inheritRight_mask_map(self):
        """
            The permissions applied.
        """
        _map = {}
        if self.lang == 'ch':
            _map = {
                '': '只有该文件夹',
                '(CI)(IO)': '只有子文件夹',
                '(OI)(IO)': '只有文件',
                '(CI)': '此文件夹和子文件夹',
                '(OI)': '此文件夹和文件',
                '(OI)(CI)(IO)': '仅子文件夹和文件',
                '(OI)(CI)': '此文件夹、子文件夹和文件'
            }

        return _map

    def get_progagateInherit_map(self):
        """
            The permission propagation.
        """
        _map = {}
        if self.lang == 'ch':
            _map = {
                '0': '不传播继承',
                '1': '传播继承'
            }

        return _map

    def get_parentInherit_map(self):
        """
             Inherit from parent node.
        """
        _map = {}
        if self.lang == 'ch':
            _map = {
                '0': '不继承',
                '1': '继承'
            }

        return _map

    def AuthsExport(self, acl_map: dict, output: str, mode: int, *args, **kwargs):

        if not acl_map or not output:
            return False

        if mode == 1:
            self.__W2Xlsx__(acl_map, output)
        elif mode == 2:
            self.__W2Xlsx2__(acl_map, output)

        return True

    def __W2Xlsx__(self, acl_map: dict, output: str, *args, **kwargs):
        """
                Write the ACL contents to the .xlsx FILE.
        :param acl_map:   {path: {'accessState': auths_list, 'subDirs': auth_subDirs, 'subFiles': auth_subFiles,'count_result': [auth_usersList]}}
        :param output:
        :param args:
        :param kwargs:
        :return:
        """
        wb = Workbook()
        ws = wb.active

        # view
        ws.views.sheetView[0].showGridLines = False
        ws.title = "Windows ACL Perm."

        # 字体
        font_consolas = Font(name="Consolas",size=11,bold=True,italic=False,color="00000000")
        font_black = Font(name="黑体",size=11,bold=True,italic=False,color="00000000")

        n_rows = 0
        n_cols = 0

        # 列表表头
        # A1,A2  B1,B2
        ws.row_dimensions[1].height = 40
        ws.row_dimensions[2].height = 20
        fille = PatternFill('solid', fgColor='00FFCC00')
        fille2 = PatternFill('solid', fgColor='00FF6600')
        for col_i, col_j in self.get_title_map().items():
            n_cols += 1
            cell_1 = ws[get_column_letter(n_cols) + str(1)]
            cell_2 = ws[get_column_letter(n_cols) + str(2)]
            # 设置样式
            cell_1.value = col_i
            cell_2.value = col_j
            cell_1.font = font_consolas
            cell_2.font = font_black
            if (n_cols < 10):
                cell_2.fill = fille
            else:
                cell_2.fill = fille2

        n_rows += 2

        # DO: CONTENTS
        for cur_k, cur_v in acl_map.items():
            if cur_v['accessState'] == '拒绝访问':
                n_rows += 1
                ws['A' + str(n_rows)] = cur_k
                for v in range(2, 26):
                    ws[get_column_letter(v) + str(n_rows)] = '拒绝访问'
                continue
            for cur_perm in cur_v["accessState"]:
                n_rows += 1
                ws['A' + str(n_rows)] = cur_k
                ws['B' + str(n_rows)] = str(cur_perm['domain'])
                ws['C' + str(n_rows)] = str(cur_perm['user'])
                ws['D' + str(n_rows)] = str(cur_perm['fullAccessMask'])
                ws['E' + str(n_rows)] = '(%s)' % str.join(',', cur_perm['accessMask'])
                ws['F' + str(n_rows)] = str(cur_perm['inherit'])
                ws['G' + str(n_rows)] = [v for k, v in self.get_parentInherit_map().items() if str(cur_perm['parentInherit']) == str(k)][0]
                ws['H' + str(n_rows)] = [v for k, v in self.get_progagateInherit_map().items() if str(cur_perm['propagateInherit']) == str(k)][0]
                ws['I' + str(n_rows)] = [v for k, v in self.get_inheritRight_int_map().items() if str(cur_perm['inheritRight']) == str(k)][0]
                ws['J' + str(n_rows)] = str(cur_perm['isAllow'])
                ws['K' + str(n_rows)] = str(cur_perm['fullControl'])
                ws['L' + str(n_rows)] = str(cur_perm['readData_listDir'])
                ws['M' + str(n_rows)] = str(cur_perm['readAttr'])
                ws['N' + str(n_rows)] = str(cur_perm['readExtAttr'])
                ws['O' + str(n_rows)] = str(cur_perm['readPermiss'])
                ws['P' + str(n_rows)] = str(cur_perm['execute_traverse'])
                ws['Q' + str(n_rows)] = str(cur_perm['writeData_addFile'])
                ws['R' + str(n_rows)] = str(cur_perm['appendData_addSubdir'])
                ws['S' + str(n_rows)] = str(cur_perm['writeAttr'])
                ws['T' + str(n_rows)] = str(cur_perm['writeExtAttr'])
                ws['U' + str(n_rows)] = str(cur_perm['delete'])
                ws['V' + str(n_rows)] = str(cur_perm['deleteChild'])
                ws['W' + str(n_rows)] = str(cur_perm['changePermiss'])
                ws['X' + str(n_rows)] = str(cur_perm['takeOwner'])
                ws['Y' + str(n_rows)] = str(cur_perm['sync'])

        # 水平垂直居中
        for row in ws['A1:Y2']:
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        side = Side(
            style=borders.BORDER_THIN,
            color=colors.BLACK,
        )

        max_cols_len_list = []

        # 设置列宽
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws_cell = ws[get_column_letter(col) + str(row)]
                ws_cell.border = Border(
                    top=side,
                    bottom=side,
                    left=side,
                    right=side
                )
                if ws_cell.value:
                    if row == 1:
                        cols_value_len_list = []  # 记录当前列所有字符串
                        for rrow in range(1, ws.max_row + 1):
                            cols_value_len_list.append(ws[get_column_letter(col) + str(rrow)].value)
                        # 计算当前列所有字符串的长度
                        count_cols_value_dict = {v: (1.2 * len(re.findall(r'[\u4e00-\u9fa5]', str(v))) + len(str(v)))
                                                 for v in cols_value_len_list}
                        # 计算当前列的最长字符串长度
                        max_count_cols_value = max(count_cols_value_dict.values())
                        max_cols_len_list.append(max_count_cols_value)

        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = max_cols_len_list[col - 1] + 2

        # 替换U+274E
        for row in range(1, ws.max_row + 1):
            for col in range(10, ws.max_column + 1):
                ws_cell = ws[get_column_letter(col) + str(row)]

                if ws_cell.value == '0' or ws_cell.value == '-1':
                    ws_cell.value = "\u274C"
                elif ws_cell.value == '1':
                    ws_cell.value = "\u2714"

        wb.save(output)
        return True

    def __W2Xlsx2__(self, acl_map: dict, output: str, *args, **kwargs):
        """
                Write the ACL contents to the .xlsx FILE.
        :param acl_map:   {path: {'accessState': auths_list, 'subDirs': auth_subDirs, 'subFiles': auth_subFiles,'count_result': [auth_usersList]}}
        :param output:
        :param args:
        :param kwargs:
        :return:
        """

        wb = Workbook()
        ws = wb.active

        user_group = []
        path_group = []

        for p, v in acl_map.items():
            if p not in path_group:
                path_group.append(p)
            for access_ in v['accessState']:
                if access_['user'] not in user_group:
                    user_group.append(access_['user'])

        # print("user_group: ", user_group)
        # print("path_group: ", path_group)

        # view
        ws.views.sheetView[0].showGridLines = False
        ws.title = "Windows ACL Perm2."
        ws.row_dimensions[2].height = 20

        # font
        font_consolas = Font(name="Consolas",size=11,bold=True,italic=False,color="00000000")
        font_black = Font(name="黑体",size=8,bold=True,italic=False,color="00000000")
        font_kai = Font(name="楷体",size=8,bold=True,italic=False,color="00000000")

        # color fill
        fille = PatternFill('solid', fgColor='00CCFFCC')
        fille2 = PatternFill('solid', fgColor='00FF0000')

        num_title_columns = len(user_group) * 2 + 2
        num_title_rows = 3

        ws.merge_cells('A1:A3')
        ws['A1'].value = 'Path'
        ws.merge_cells('B1:B3')
        ws['B1'].value = 'Permission'
        ws.merge_cells('{}1:{}1'.format(get_column_letter(3),get_column_letter(num_title_columns)))
        ws['C1'].value = 'User Group'

        # col 3
        j = 1
        for i in range(len(user_group)):
            c = get_column_letter(j+2)
            r = get_column_letter(j+3)
            j = j + 2
            ws.merge_cells('{}2:{}2'.format(c,r))
            ws[c+'2'].value = str(user_group[i])
            ws[c+'3'].value = '授予'
            ws[r+'3'].value = '拒绝'

        # row 4
        # 保存需要填入信息的
        wrap_true_list = [2]
        j = 4
        for i in range(len(path_group)):
            e = j + 16
            ws.merge_cells('A{}:A{}'.format(j,e))
            ws['A' + str(j)].value = path_group[i]

            # for all columns
            # for l in range(2,num_title_columns-1):
            #     col_index = get_column_letter(l)
            #     ws.merge_cells('{}{}:{}{}'.format(col_index,j,col_index,e))

            # 权限信息详细
            item = 0
            for k in range(j,e):
                ws['B' + str(k)].value = list(self.get_per_map().values())[item]
                item += 1

            # 填入具体信息
            vi = acl_map[path_group[i]]
            if vi['accessState'] == '拒绝访问':
                # 该路径无法获取任何权限信息
                pass
            else:
                for perm in vi['accessState']:
                    usr_idx  = user_group.index(perm['user'])
                    col_idx = get_column_letter(usr_idx*2+3)
                    col_idx2 = get_column_letter(usr_idx*2+3+1)

                    self.__fill_ws__(perm, col_idx, col_idx2, ws, 'isAllow', j)
                    self.__fill_ws__(perm, col_idx, col_idx2, ws, 'fullControl', j+1)
                    self.__fill_ws__(perm, col_idx, col_idx2, ws, 'readData_listDir', j+2)
                    self.__fill_ws__(perm, col_idx, col_idx2, ws, 'readAttr', j+3)
                    self.__fill_ws__(perm, col_idx, col_idx2, ws, 'readExtAttr', j+4)
                    self.__fill_ws__(perm, col_idx, col_idx2, ws, 'readPermiss', j+5)
                    self.__fill_ws__(perm, col_idx, col_idx2, ws, 'execute_traverse', j+6)
                    self.__fill_ws__(perm, col_idx, col_idx2, ws, 'writeData_addFile', j+7)
                    self.__fill_ws__(perm, col_idx, col_idx2, ws, 'appendData_addSubdir', j+8)
                    self.__fill_ws__(perm, col_idx, col_idx2, ws, 'writeAttr', j+9)
                    self.__fill_ws__(perm, col_idx, col_idx2, ws, 'writeExtAttr', j+10)
                    self.__fill_ws__(perm, col_idx, col_idx2, ws, 'delete', j+11)
                    self.__fill_ws__(perm, col_idx, col_idx2, ws, 'deleteChild', j+12)
                    self.__fill_ws__(perm, col_idx, col_idx2, ws, 'changePermiss', j+13)
                    self.__fill_ws__(perm, col_idx, col_idx2, ws, 'takeOwner', j+14)
                    self.__fill_ws__(perm, col_idx, col_idx2, ws, 'sync', j+15)

                    authorized_ = ""
                    refuse_ = ""
                    temp_ = ""

                    if perm['parentInherit'] == 1:
                        temp_ += "\n* 从父对象继承"
                    else:
                        temp_ += "\n* 不从父对象继承"

                    if perm['propagateInherit'] == 1:
                        temp_ += "\n* 传播继承"
                    else:
                        temp_ += "\n* 不传播继承"

                    temp_ += "\n* " + [v for k, v in self.get_inheritRight_int_map().items() if str(perm['inheritRight']) == str(k)][0]

                    if "(DENY)" not in perm['accessMask'] and "(N)" not in perm['accessMask']:
                        authorized_ += '(%s)' % str.join(',', perm['accessMask'])
                        authorized_ += temp_
                    else:
                        refuse_ += '(%s)' % str.join(',', perm['accessMask'])
                        refuse_ += temp_

                    ws[col_idx + str(j+16)].value = authorized_
                    ws[col_idx2 + str(j+16)].value = refuse_
                    ws[col_idx + str(j+16)].font = font_black
                    ws[col_idx2 + str(j+16)].font = font_kai

                    if refuse_ != "":
                        ws[col_idx2 + str(j+16)].fill = fille2
                    elif authorized_ != "":
                        ws[col_idx + str(j+16)].fill = fille

                    ws.column_dimensions[col_idx].width = 10
                    ws.column_dimensions[col_idx2].width = 10
                    ws.row_dimensions[j+16].height = 88

                    wrap_true_list.append(j+16)

            j = j + 17

        all_rows = len(path_group) * 17 + 3

        side = Side(
            style=borders.BORDER_THIN,
            color=colors.BLACK,
        )

        # 水平垂直居中
        for row in ws['A1:{}{}'.format(get_column_letter(num_title_columns),all_rows)]:
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    top=side,
                    bottom=side,
                    left=side,
                    right=side
                )

        # 设置行自动换行
        for row_numb in wrap_true_list:
            for cell in ws[row_numb]:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for i in range(4, all_rows):
            ws['A'+str(i)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 冻结窗口
        ws.freeze_panes = "C4"

        # 设置列宽
        # 计算当前列所有字符串的长度
        #cols_len = [len(v) for v in path_group]
        #max_column = max(cols_len)
        #max_col_len = max_column * 2
        ws.column_dimensions[get_column_letter(1)].width = 12 + 2
        ws.column_dimensions[get_column_letter(2)].width = 20 + 2

        wb.save(output)
        return True

    def __fill_ws__(self, perm: list, col_idx: str, col_idx2: str, ws: Workbook, widx: str, rowj: int):
        color_fill = PatternFill('solid', fgColor='00CCFFCC')
        color_fill2 = PatternFill('solid', fgColor='00CCFFFF')
        if perm[widx] == 1:
            ws['{}'.format(col_idx) + str(rowj)] = "\u2714"
            ws['{}'.format(col_idx) + str(rowj)].fill = color_fill
        elif perm[widx] == 0:
            ws['{}'.format(col_idx2) + str(rowj)] = "\u2714"
            ws['{}'.format(col_idx2) + str(rowj)].fill = color_fill2
        return True